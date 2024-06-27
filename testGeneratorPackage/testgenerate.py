import openpyxl as xl
from pathlib import Path
import os

from testGeneratorPackage import utils


def generate_test_plan(dictionary_requirement, file_path, test_item_sheet_name, test_team_sheet_name):
    a = os.path.split(os.getcwd())[0]

    path = Path(a + file_path)

    # Load the workbook
    wb = xl.load_workbook(path)
    # dictionary with all needed items only
    dictionary_to_filter = {}
    for key in dictionary_requirement:
        if dictionary_requirement[key] is not None:
            dictionary_to_filter[key] = dictionary_requirement[key]

    # Try to get the result sheet, if not exist create one
    try:
        result_sheet = wb['Result_Sheet']
        wb.__delitem__('Result_Sheet')
    except KeyError:
        pass
    finally:
        wb.create_sheet('Result_Sheet', 1)
        wb.save(path)
        result_sheet = wb['Result_Sheet']

    input_sheet = wb[test_item_sheet_name]
    input_sheet_team = wb[test_team_sheet_name]
    number_of_rows = input_sheet.max_row
    dictionary_index = utils.create_index_dictionary(input_sheet)
    column_number = len(dictionary_index)
    result_row = 1

    # Iterate through rows in the input sheet
    for i in range(1, number_of_rows + 1):
        not_to_be_added = 0

        # Check if row matches the requirements
        if (i != 1):
            for j in dictionary_to_filter:
                if input_sheet.cell(i, dictionary_index[j]).value != dictionary_to_filter[j]:
                    not_to_be_added += 1

        # If row matches, add it to the result sheet
        if not_to_be_added == 0:
            for start in range(1, column_number + 1):
                result_sheet.cell(result_row, start).value = input_sheet.cell(i, start).value
                print(input_sheet.cell(i, start).value)
            result_row += 1

    # Save the workbook
    wb.save(path)
